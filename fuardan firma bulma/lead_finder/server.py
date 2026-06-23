import json
import os
import urllib.parse
from http.server import SimpleHTTPRequestHandler, HTTPServer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PORT = 8500
EXCEL_PATH = r"C:\Users\GilboTeknik\Desktop\SKANOPT OÖC\fuardan firma bulma\Robotsepeti_Tum_Fuar_Listesi_BirlestikOtomatik_Kurtarıldı.xlsx"
JSON_PATH = "leads.json"

class LeadFinderHandler(SimpleHTTPRequestHandler):
    def end_headers(self):
        # Allow CORS and caching preventions for development convenience
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_POST(self):
        if self.path == '/api/save':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                leads = json.loads(post_data.decode('utf-8'))
                
                # Save to leads.json
                with open(JSON_PATH, 'w', encoding='utf-8') as f:
                    json.dump(leads, f, ensure_ascii=False, indent=4)
                
                # Save to leads.js fallback as well
                js_path = JSON_PATH.replace('.json', '.js')
                with open(js_path, 'w', encoding='utf-8') as f:
                    f.write("window.LEADS_DATA = ")
                    json.dump(leads, f, ensure_ascii=False, indent=4)
                    f.write(";")
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                response = {"status": "success", "message": "Firma verileri leads.json dosyasına başarıyla kaydedildi!"}
                self.wfile.write(json.dumps(response).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                response = {"status": "error", "message": f"Kaydedilemedi: {str(e)}"}
                self.wfile.write(json.dumps(response).encode('utf-8'))

        elif self.path == '/api/export_excel':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                payload = json.loads(post_data.decode('utf-8'))
                leads = payload.get('leads', [])
                
                if not os.path.exists(EXCEL_PATH):
                    # If the file doesn't exist, create a blank workbook
                    wb = openpyxl.Workbook()
                    wb.save(EXCEL_PATH)
                
                # Load workbook
                wb = openpyxl.load_workbook(EXCEL_PATH)
                
                # Check and delete existing "Hedef Firmalar" sheet to overwrite
                sheet_name = "Hedef Firmalar"
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                # Create a new sheet
                ws = wb.create_sheet(title=sheet_name)
                ws.views.sheetView[0].showGridLines = True
                
                # Columns: B to H
                # We start column index from 2 (B)
                headers = ["firma ismi", "ülke", "sektör", "fuar", "web sitesi linki", "iletişim bilgileri", "not"]
                
                # Styling definitions
                green_header_fill = PatternFill(start_color="165B33", end_color="165B33", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                data_font = Font(name="Calibri", size=11, color="000000")
                link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
                
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                data_align = Alignment(horizontal="left", vertical="center")
                not_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                thin_side = Side(border_style="thin", color="D3D3D3")
                border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
                # Alternate rows colors
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                grey_fill = PatternFill(start_color="F2F4F2", end_color="F2F4F2", fill_type="solid")
                
                # Set row heights
                ws.row_dimensions[1].height = 15 # Empty row 1
                ws.row_dimensions[2].height = 28 # Header row
                
                # Write headers starting at B2 (column 2, row 2)
                for col_idx, header_text in enumerate(headers, start=2):
                    cell = ws.cell(row=2, column=col_idx, value=header_text)
                    cell.fill = green_header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = border_all
                    
                # Write data starting at row 3
                for row_idx, lead in enumerate(leads, start=3):
                    ws.row_dimensions[row_idx].height = 24
                    
                    row_fill = grey_fill if row_idx % 2 == 0 else white_fill
                    
                    # Columns map:
                    # B: firma ismi
                    # C: ülke
                    # D: sektör
                    # E: fuar
                    # F: web sitesi linki
                    # G: iletişim bilgileri
                    # H: not
                    
                    # Col B
                    cell_name = ws.cell(row=row_idx, column=2, value=lead.get('firma_ismi', ''))
                    # Col C
                    cell_country = ws.cell(row=row_idx, column=3, value=lead.get('ulke', ''))
                    # Col D
                    cell_sector = ws.cell(row=row_idx, column=4, value=lead.get('sektor', ''))
                    # Col E
                    cell_fair = ws.cell(row=row_idx, column=5, value=lead.get('fuar', ''))
                    
                    # Col F: web sitesi linki as clickable hyperlink
                    website = lead.get('web_sitesi_linki', '')
                    cell_web = ws.cell(row=row_idx, column=6)
                    if website:
                        if not website.startswith('http'):
                            target_url = "http://" + website
                        else:
                            target_url = website
                        cell_web.value = f'=HYPERLINK("{target_url}", "{website}")'
                        cell_web.font = link_font
                    else:
                        cell_web.value = ""
                        cell_web.font = data_font
                        
                    # Col G
                    cell_contact = ws.cell(row=row_idx, column=7, value=lead.get('iletisim_bilgileri', ''))
                    # Col H
                    cell_note = ws.cell(row=row_idx, column=8, value=lead.get('not', ''))
                    
                    # Apply general styling to all cells in the row
                    for col in range(2, 9):
                        c = ws.cell(row=row_idx, column=col)
                        c.fill = row_fill
                        if col != 6: # Web link font already set
                            c.font = data_font
                        c.border = border_all
                        if col == 8: # Col H (not) wraps text
                            c.alignment = not_align
                        else:
                            c.alignment = data_align
                            
                # Auto-fit column widths
                # For Col H (not), we set a fixed width of 60 since it contains long text
                # Column A is kept empty/narrow
                ws.column_dimensions['A'].width = 3
                ws.column_dimensions['B'].width = 32 # firma ismi
                ws.column_dimensions['C'].width = 12 # ülke
                ws.column_dimensions['D'].width = 25 # sektör
                ws.column_dimensions['E'].width = 20 # fuar
                ws.column_dimensions['F'].width = 30 # web sitesi
                ws.column_dimensions['G'].width = 45 # iletişim bilgileri
                ws.column_dimensions['H'].width = 65 # not (wrap text active)
                
                # Save the workbook
                wb.save(EXCEL_PATH)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                response = {"status": "success", "message": f"Excel başarıyla güncellendi! '{sheet_name}' sayfası eklendi/güncellendi."}
                self.wfile.write(json.dumps(response).encode('utf-8'))
            except OSError as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                if e.errno == 13 or "permission denied" in str(e).lower() or "denied" in str(e).lower():
                    msg = "Hata: Excel dosyası şu anda başka bir programda (örneğin Microsoft Excel) açık. Lütfen dosyayı kapatıp tekrar deneyin!"
                else:
                    msg = f"Excel işletim sistemi hatası: {str(e)}"
                response = {"status": "error", "message": msg}
                self.wfile.write(json.dumps(response).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                response = {"status": "error", "message": f"Excel yazma hatası: {str(e)}"}
                self.wfile.write(json.dumps(response).encode('utf-8'))
        else:
            super().do_POST()

def run(server_class=HTTPServer, handler_class=LeadFinderHandler):
    server_address = ('', PORT)
    httpd = server_class(server_address, handler_class)
    print(f"\n=======================================================")
    print(f"SKANOPT Fuar Firma Arama & B2B CRM Portalı Başlatıldı!")
    print(f"Erişim Adresi: http://localhost:{PORT}")
    print(f"Geliştirici Notu: Mevcut landing sayfasını bozmadan çalışır.")
    print(f"Kapatmak için: Terminalde Ctrl+C tuşlarına basın.")
    print(f"=======================================================\n")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nSunucu durduruluyor...")
        httpd.server_close()

if __name__ == '__main__':
    run()
